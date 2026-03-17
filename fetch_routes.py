"""
RPA - Ametller Origen: Extracción de rutas diarias → Excel
- Stop empieza en 0
- Paginación completa via intercepción de todas las páginas de la API
- 2 tiendas x 2 días = hasta 4 Excel
"""

import os
import json
import asyncio
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
PORTAL_EMAIL    = os.environ["PORTAL_EMAIL"]
PORTAL_PASSWORD = os.environ["PORTAL_PASSWORD"]

PORTAL_URL = "https://control.instaleap.io"
CLIENT_ID  = "AMETLLER_ORIGEN"
STATES     = ["CREATED", "ON_BOARDING", "PROCESSING"]
PAGE_LIMIT = 10  # El portal usa limit=10 por defecto

STORES = [
    {"id": "70fcffac-e7de-44ca-845b-f316fd5b874e", "name": "ElPrat"},
    {"id": "6258d460-6aad-40e2-9de8-6914d4ea7a96", "name": "Garraf"},
]


# ─────────────────────────────────────────────
# 1. FETCH TODAS LAS PÁGINAS DE UNA TIENDA/FECHA
# ─────────────────────────────────────────────
async def fetch_routes_for_store_date(page, store_id: str, store_name: str, target_date) -> list:
    all_routes  = []
    total_pages = 1
    pages_data  = {}  # offset -> routes

    # ── Interceptor: captura TODAS las respuestas de /nebula/routing ──
    async def handle_response(response):
        nonlocal total_pages
        if "nebula/routing" in response.url and response.status == 200:
            try:
                data   = await response.json()
                routes = data.get("routes", [])
                tp     = data.get("total_pages", 1)
                total_pages = max(total_pages, tp)

                # Extraer offset de la URL
                offset = 0
                if "offset=" in response.url:
                    try:
                        offset = int(response.url.split("offset=")[1].split("&")[0])
                    except:
                        pass

                if offset not in pages_data:
                    pages_data[offset] = routes
                    print(f"    📥 [{store_name}] offset={offset}: {len(routes)} rutas (total_pages={tp})")

            except Exception as e:
                print(f"    ⚠️  Error parseando: {e}")

    page.on("response", handle_response)

    # ── Construir URL base de la API ──
    madrid   = timezone(timedelta(hours=1))
    from_dt  = datetime(target_date.year, target_date.month, target_date.day, 0, 0, 0,
                        tzinfo=madrid)
    to_dt    = datetime(target_date.year, target_date.month, target_date.day, 23, 59, 59,
                        tzinfo=madrid)
    from_str = from_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    to_str   = to_dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.999Z")

    states_qs = "".join([f"&states[]={s}" for s in STATES])
    api_base  = (
        f"https://avt-backend.instaleap.io/nebula/routing"
        f"?store_id={store_id}"
        f"&client_id={CLIENT_ID}"
        f"&limit={PAGE_LIMIT}"
        f"&from={from_str}"
        f"&to={to_str}"
        f"{states_qs}"
    )

    date_str   = target_date.strftime("%Y-%m-%d")
    portal_url = (
        f"{PORTAL_URL}/routes"
        f"?storeId={store_id}"
        f"&date={date_str}"
        f"&status=CREATED,ON_BOARDING,PROCESSING"
    )

    # ── Cargar la página del portal (esto triggerea offset=0) ──
    print(f"    🔄 [{store_name}] Navegando a {date_str}...")
    await page.goto(portal_url, wait_until="networkidle", timeout=30000)
    await page.wait_for_timeout(3000)

    # ── Esperar a tener la página 0 ──
    for _ in range(10):
        if 0 in pages_data:
            break
        await page.wait_for_timeout(1000)

    if 0 not in pages_data:
        print(f"    ⚠️  No se recibió página 0 para {store_name} {date_str}")
        page.remove_listener("response", handle_response)
        return []

    # ── Pedir páginas adicionales directamente desde el navegador ──
    for page_num in range(1, total_pages):
        offset   = page_num * PAGE_LIMIT
        api_url  = f"{api_base}&offset={offset}"

        print(f"    📡 Pidiendo página {page_num+1}/{total_pages} (offset={offset})...")

        # Usar fetch() desde dentro del navegador (misma sesión/cookies)
        result = await page.evaluate(f"""
            async () => {{
                try {{
                    const r = await fetch("{api_url}", {{
                        method: "GET",
                        headers: {{ "accept": "application/json", "content-type": "application/json" }},
                        credentials: "include"
                    }});
                    return {{ status: r.status, text: await r.text() }};
                }} catch(e) {{
                    return {{ status: 0, text: e.toString() }};
                }}
            }}
        """)

        if result["status"] == 200:
            try:
                data   = json.loads(result["text"])
                routes = data.get("routes", [])
                if offset not in pages_data:
                    pages_data[offset] = routes
                    print(f"    📥 Página {page_num+1}: {len(routes)} rutas adicionales")
            except Exception as e:
                print(f"    ⚠️  Error parseando página {page_num+1}: {e}")
        else:
            print(f"    ⚠️  Error página {page_num+1}: status={result['status']} {result['text'][:100]}")

        await page.wait_for_timeout(500)

    page.remove_listener("response", handle_response)

    # ── Combinar todas las páginas en orden ──
    for offset in sorted(pages_data.keys()):
        all_routes.extend(pages_data[offset])

    print(f"    ✅ Total [{store_name}] {date_str}: {len(all_routes)} rutas")
    return all_routes


# ─────────────────────────────────────────────
# 2. LOGIN + FETCH TODAS LAS COMBINACIONES
# ─────────────────────────────────────────────
async def fetch_all_routes() -> list:
    print("🔐 Iniciando login con Playwright...")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            viewport={"width": 1280, "height": 900}
        )
        page = await context.new_page()

        # ── Login ──
        await page.goto(PORTAL_URL, wait_until="networkidle", timeout=30000)
        print("📄 Portal cargado")
        await page.wait_for_timeout(2000)

        await page.wait_for_selector('#email', timeout=15000)
        await page.fill('#email', PORTAL_EMAIL)
        print("✉️  Email introducido")
        await page.click('button:has-text("Continue")')

        await page.wait_for_selector('input[type="password"]', timeout=15000)
        await page.fill('input[type="password"]', PORTAL_PASSWORD)
        print("🔑 Contraseña introducida")
        await page.click('button[type="submit"], button:has-text("Log in"), button:has-text("Continue"), button:has-text("Sign in")')

        await page.wait_for_url("**/routes**", timeout=30000)
        print("✅ Login exitoso\n")
        await page.wait_for_timeout(3000)

        # ── Fechas ──
        madrid   = timezone(timedelta(hours=1))
        today    = datetime.now(madrid).date()
        tomorrow = today + timedelta(days=1)

        results = []

        for store in STORES:
            for date_obj, date_label in [(today, "HOY"), (tomorrow, "MAÑANA")]:
                print(f"\n📅 [{store['name']}] {date_label} ({date_obj})...")
                routes = await fetch_routes_for_store_date(
                    page, store["id"], store["name"], date_obj
                )
                filename = f"rutas_{store['name']}_{date_obj.strftime('%Y-%m-%d')}.xlsx"
                results.append({
                    "store_name": store["name"],
                    "date":       date_obj,
                    "date_label": date_label,
                    "routes":     routes,
                    "filename":   filename,
                })
                await page.wait_for_timeout(1000)

        await browser.close()

    return results


# ─────────────────────────────────────────────
# 3. GENERAR EXCEL (Stop empieza en 0)
# ─────────────────────────────────────────────
def parse_dt(s: str) -> str:
    dt = datetime.strptime(s, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone(timedelta(hours=1))).strftime("%d/%m/%Y %H:%M")


def generate_excel(routes: list, output_path: str, store_name: str, date_label: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Routes"

    header_fill  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_light   = PatternFill("solid", start_color="EBF0FA", end_color="EBF0FA")
    fill_white   = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")

    col_headers = ["Route Number", "Job Number", "Stop", "Delivery From", "Delivery To"]
    col_widths  = [28, 28, 8, 22, 22]

    for col, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    row = 2
    for route in routes:
        route_num = route.get("route_number", "")
        for stop_idx, task in enumerate(route.get("tasks", []), 0):  # ✅ empieza en 0
            fill = fill_light if row % 2 == 0 else fill_white
            values = [
                route_num,
                task.get("job_number", ""),
                stop_idx,
                parse_dt(task["from"]) if task.get("from") else "",
                parse_dt(task["to"])   if task.get("to")   else "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font      = Font(name="Arial", size=10)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = Alignment(
                    horizontal="center" if col == 3 else "left",
                    vertical="center"
                )
            row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{row - 1}"
    wb.save(output_path)
    print(f"  ✅ Excel generado: {output_path} ({row - 2} filas)")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
async def main():
    results   = await fetch_all_routes()
    generated = []

    for entry in results:
        if entry["routes"]:
            generate_excel(
                entry["routes"],
                entry["filename"],
                entry["store_name"],
                entry["date_label"],
            )
            generated.append(entry["filename"])
        else:
            print(f"⚠️  Sin rutas para {entry['store_name']} {entry['date_label']} — no se genera Excel.")

    if generated:
        print(f"\n📊 Archivos generados: {generated}")
    else:
        print("\n⚠️  No se generó ningún Excel.")


if __name__ == "__main__":
    asyncio.run(main())
